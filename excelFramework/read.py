import openpyxl

book = openpyxl.load_workbook(filename="F:\SeleniumComponents\cricket.xlsx")
sheet = book.active
row = sheet.max_row
column = sheet.max_column
print("row count",row)
print("column count",column)

for r in range(1,row+1):
    for c in range(1,column+1):
        print(sheet.cell(row=r,column=c).value,end="\t\t")
    print()
book.close()