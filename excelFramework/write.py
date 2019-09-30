import openpyxl

book = openpyxl.load_workbook(filename="F:\SeleniumComponents\cricket.xlsx")
sheet = book.active
row = sheet.max_row
column = sheet.max_column
print("row count",row)
print("column count",column)
sheet.cell(row=7, column=1).value = "Raina"
sheet.cell(row=7, column=2).value = 16
sheet.cell(row=7, column=3).value = 120000
book.save(filename="F:\SeleniumComponents\cricket.xlsx")
book.close()