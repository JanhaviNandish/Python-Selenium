import openpyxl
from selenium import webdriver

book = openpyxl.load_workbook(filename="F:\SeleniumComponents\magento.xlsx")
sheet = book.active
row = sheet.max_row
column = sheet.max_column
print("row count",row)
print("column count",column)
for r in range(2,row+1):
    action = sheet.cell(row=r, column=3).value
    print(action)

    if action=='open':
        chrome = webdriver.Chrome(executable_path='F:\SeleniumComponents\chromedriver.exe')
        chrome.maximize_window()
        chrome.implicitly_wait(time_to_wait=20)
    elif action == 'navigate':
        chrome.get(url=sheet.cell(row=r,column=5).value)
    elif action == 'click':
        chrome.find_element_by_xpath(xpath=sheet.cell(row=r,column=4).value).click()
    elif action == 'type':
        chrome.find_element_by_xpath(xpath=sheet.cell(row=r,column=4).value).send_keys(sheet.cell(row=r,column=5).value)
    elif action == 'quit':
        chrome.quit()
    else:
        print("invalid action")

